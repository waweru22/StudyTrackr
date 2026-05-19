import re
from datetime import time as dt_time

import openpyxl

from app import db
from app.models.timetable_entry import TimetableEntry
from app.models.user import User

LEVEL_TO_SHEET = {
    100: '100-level',
    200: '200-level',
    300: '300-level',
    400: '400-level',
}


def user_has_timetable(user_id: int) -> bool:
    return (
        db.session.query(TimetableEntry.id)
        .filter_by(user_id=user_id)
        .limit(1)
        .first()
        is not None
    )


def has_schedule_blocks(user_id: int) -> bool:
    from app.models.session import ScheduleBlock

    return (
        db.session.query(ScheduleBlock.id)
        .filter_by(user_id=user_id)
        .limit(1)
        .first()
        is not None
    )


def ensure_timetable_flag(user_id: int) -> bool:
    """Lightweight flag sync (no schedule deletes on read)."""
    user = User.query.get(user_id)
    if not user:
        return False

    has_timetable = user_has_timetable(user_id)
    if user.timetable_uploaded != has_timetable:
        user.timetable_uploaded = has_timetable
        db.session.commit()
    return has_timetable


def normalise_code(code: str) -> str:
    if not code:
        return ''
    return str(code).replace(' ', '').strip().upper()


def parse_time_string(time_str: str):
    '''
    Returns (start: dt_time, end: dt_time) or (None, None)
    if parsing fails.
    '''
    if not time_str:
        return None, None

    s = time_str.strip().upper().replace(' ', '')

    pattern = r'^(\d{1,2})(?::\d{2})?(AM|PM)-(\d{1,2})(?::\d{2})?(AM|PM)$'
    match = re.match(pattern, s)

    if not match:
        return None, None

    start_h = int(match.group(1))
    start_period = match.group(2)
    end_h = int(match.group(3))
    end_period = match.group(4)

    def to_24h(h, period):
        if period == 'AM':
            return 0 if h == 12 else h
        else:
            return h if h == 12 else h + 12

    start_24 = to_24h(start_h, start_period)
    end_24 = to_24h(end_h, end_period)

    if end_24 < start_24:
        end_24 = to_24h(end_h, start_period)

    if not (0 <= start_24 <= 23 and 0 <= end_24 <= 23):
        return None, None
    if end_24 <= start_24:
        return None, None

    return dt_time(start_24, 0), dt_time(end_24, 0)


def parse_timetable(file_path: str, user_id: int) -> dict:
    '''
    Main entry point. Parses the Excel file for the user's level,
    extracts their class slots, saves to DB, marks timetable
    as uploaded.
    '''
    try:
        user = User.query.get(user_id)
        if not user:
            return {"success": False, "error": "User not found"}

        if user_has_timetable(user_id):
            return {
                "success": False,
                "error": "Timetable already uploaded. Your schedule cannot be re-uploaded.",
            }

        level = user.level
        sheet_name = LEVEL_TO_SHEET.get(level)
        if not sheet_name:
            return {
                "success": False,
                "error": f"No timetable sheet for level {level}"
            }

        wb = openpyxl.load_workbook(file_path, read_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' not found in file"
                }

            ws = wb[sheet_name]

            enrolled_codes = {
                normalise_code(c.code)
                for c in user.courses
            }

            TimetableEntry.query.filter_by(user_id=user_id).delete()

            entries_saved = 0
            skipped = 0
            classes = []
            row_index = 0

            for row in ws.iter_rows(values_only=True):
                row_index += 1

                if row_index <= 2:
                    continue

                if not row or all(cell is None for cell in row):
                    continue

                if row[0] and str(row[0]).startswith('Column'):
                    continue

                raw_code = row[0]
                raw_name = row[1]
                raw_time = row[2]
                raw_day = row[3]
                raw_venue = row[4] if len(row) > 4 else None

                if not raw_code or not raw_time or not raw_day:
                    skipped += 1
                    continue

                code = normalise_code(str(raw_code))
                day = str(raw_day).strip().capitalize()

                valid_days = [
                    'Monday', 'Tuesday', 'Wednesday',
                    'Thursday', 'Friday', 'Saturday', 'Sunday'
                ]
                if day not in valid_days:
                    skipped += 1
                    continue

                start_t, end_t = parse_time_string(str(raw_time))
                if start_t is None or end_t is None:
                    print(
                        f'[Timetable] Skipping malformed time '
                        f'in row {row_index}: {raw_time}'
                    )
                    skipped += 1
                    continue

                name_str = str(raw_name) if raw_name else ''
                section = None
                sec_match = re.search(
                    r'\(Section\s*(\d+)\)', name_str,
                    re.IGNORECASE
                )
                if sec_match:
                    section = f"Section {sec_match.group(1)}"
                    clean_name = re.sub(
                        r'\s*\(Section\s*\d+\)', '', name_str
                    ).strip()
                else:
                    clean_name = name_str.strip()

                venue_str = str(raw_venue).strip() if raw_venue else None

                if code not in enrolled_codes:
                    skipped += 1
                    continue

                entry = TimetableEntry(
                    user_id=user_id,
                    course_code=code,
                    course_name=clean_name,
                    day_of_week=day,
                    start_time=start_t,
                    end_time=end_t,
                    venue=venue_str,
                    section=section
                )
                db.session.add(entry)
                entries_saved += 1

                classes.append({
                    "course_code": code,
                    "course_name": clean_name,
                    "day": day,
                    "start_time": start_t.strftime('%H:%M'),
                    "end_time": end_t.strftime('%H:%M'),
                    "venue": venue_str,
                    "section": section
                })

            db.session.commit()

            user.timetable_uploaded = entries_saved > 0
            db.session.commit()

            print(
                f'[Timetable] Saved {entries_saved} entries, '
                f'skipped {skipped} rows for user {user_id}'
            )

            return {
                "success": True,
                "entries_saved": entries_saved,
                "skipped": skipped,
                "classes": classes,
                "error": None
            }
        finally:
            wb.close()

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return {
            "success": False,
            "error": str(e),
            "entries_saved": 0,
            "skipped": 0,
            "classes": []
        }
